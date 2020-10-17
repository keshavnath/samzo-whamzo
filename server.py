from flask import Flask, render_template, request, redirect, url_for
import csv
import datetime
from openpyxl import Workbook, load_workbook
app=Flask(__name__)

answers=["funeral blues", "coaster", "halwa", "chudail", "ghutne"]
blocked=["submit", "go-ascii", "redirect-home"]

@app.route("/")
def home_page():
    return render_template("index.html")

def username_in_xl(username):
    wb=load_workbook("database.xlsx")
    ws=wb.active
    i=1
    while i>0 and ws.cell(row=i, column=1).value!=None:
        if ws.cell(row=i, column=1).value==username:
            data={"username":username, "level":ws.cell(row=i, column=2).value}
            wb.save("database.xlsx")
            return (data)
        i=i+1    
    ws.cell(row=i, column=1).value=username
    ws.cell(row=i, column=2).value=0
    data={"username":username, "level":0}
    wb.save("database.xlsx")
    return (data)

'''
def hall_of_fame(player):
    username=player["username"]
    level=player["level"]
    wb=load_workbook("hall_of_fame.xlsx")
    ws=wb.active
    i=1
    while i>0 and ws.cell(row=i, column=1).value!=None:
        ws.cell(row=i, column=1).value=username
        ws.cell(row=i, column=2).value=(datetime.datetime.now()).strftime("%x")
'''

def update_xl(player):
    #if(int(player["level"])==5):
    #    hall_of_fame(player)
    username=player["username"]
    level=int(player["level"])
    wb=load_workbook("database.xlsx")
    ws=wb.active
    i=1
    while i>0 and ws.cell(row=i, column=1).value!=None:
        if ws.cell(row=i, column=1).value==username:
            ws.cell(row=i, column=2).value=level
            wb.save("database.xlsx")
            break
        i=i+1


@app.route("/<string:username>", methods=["POST", "GET"])
def user_page(username):
    player_info=username_in_xl(username)
    level=int(player_info["level"])

    if request.method=="POST":
        user_ans=str(request.form.to_dict()["answer"])
        if user_ans.lower()==answers[level]:
            player_info["level"]=int(player_info["level"])+1
            level=level+1
            update_xl(player_info)
            message="Previous Answer Was Correct!"
        else:
            message="Incorrect Answer!"

        return render_template(f"level{level}.html", username=username,message=message)
    else:
        #print("lmao")
        return render_template(f"level{level}.html", username=username,message="")


@app.route("/submit", methods=["POST", "GET"])
def submit():
    if request.method=="POST":
        data = request.form.to_dict()
        '''
        player_info=username_in_csv(data)
        username=player_info["username"]
        print(player_info)
        '''
        username=str(data["username"])
        if(username.isalnum()==True):
            return redirect(url_for("user_page", username=username))
        else:
            return render_template("error.html", message="Your username was invalid. Please try again.")
    else:
        return render_template("error.html", message="You do not have permission to access this resource.")

'''
@app.route("/thankyou")
def page():
    return render_template("thankyou.html")
'''

@app.route("/redirect-home", methods=["POST", "GET"])
def go_home():
    return redirect("/")

@app.route("/go-ascii", methods=["POST", "GET"])
def go_ascii():
    return render_template("ascii.html")

'''
def username_in_csv(data):
    with open("database.csv", mode='r', newline="") as database:
        username=str(data["username"]).lower()
        csv_checker=csv.DictReader(database, delimiter=",", quotechar="'", quoting=csv.QUOTE_MINIMAL)
        flag=0
        level=0
        for line in csv_checker:
            if line["username"]==username:
                level=line["level"]
                flag=1
                break
        
        if (flag==0):
            with open("database.csv", mode='a', newline="") as database2:
                csv_writer=csv.writer(database2, delimiter=",", quotechar="'", quoting=csv.QUOTE_MINIMAL)
                csv_writer.writerow([username,0])

        return ({"username":username, "level":level})
'''

'''
def update_csv(player):
    with open("database.csv", mode="r+", newline="") as database3:
        csv_updater=csv.DictReader(database3, delimiter=",", quotechar="'", quoting=csv.QUOTE_MINIMAL)
        csv_writer=csv.writer(database3, delimiter=",", quotechar="'", quoting=csv.QUOTE_MINIMAL)
        username=player["username"]
        level=player["level"]
        for row in csv_updater:
            if row["username"]==username:
                csv_writer.writerow([username,level])
'''


'''
def update_csv(player):
    username=player["username"]
    level=player["level"]
    database=pd.read_csv("./database.csv")
    to_add=pd.DataFrame([username, level], columns=["username","level"])
'''

'''
@app.route("/answer", methods=["POST", "GET"])
def check_answer():
    if request.method=="POST":
        user_ans=request.form.to_dict()["answer"]
'''


'''
@app.route("/play")
def play():
    global player_info
    if(player_info==0):
        return render_template("error.html")
    else:
        name=player_info["username"]
        return redirect(url_for(f"{name}", username=name))
'''